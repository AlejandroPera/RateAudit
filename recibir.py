import win32com.client as win32
import datetime
import pandas as pd
import xlrd
import time
from glob import glob 
import schedule
import re
from openpyxl import load_workbook

path='D:\\Trabajo\\Test\\'            #direccion donde se guarda el csv creado
arrFiles=glob(path+"*.xlsx")
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts
print(arrFiles)
SubjectsArray=[]
repetidosArr=[]
numRepetidos=[]
arrComplexData=[]


def succesfullRetrieval():                                 #Funcion que clasifica aceptados y rechazados
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)                         # "6" refers to the index of a folder - in this case the inbox.                                      
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)
    if re.findall('~',arrFiles[-1]):
        fname=arrFiles[-2]
    else:
        fname=arrFiles[-1]
    workbook = xlrd.open_workbook(fname, on_demand=True)        #Determina el numero de filas
    sheet=workbook.sheet_by_name('Sheet1')
    row_count=sheet.nrows                                    #Lee el numero de filas
    workbook.release_resources()
    baseDatos = pd.read_excel(fname)  
    xfile = load_workbook(fname)
    sheetName = xfile.get_sheet_by_name('Sheet1')

    

    for message in messages:                                        #Hace un barrido por cada correo en bandeja
        subject_content = message.Subject
        subjectSplit=subject_content.split()                      #Separa el asunto del correo por espacios
        if (len(subjectSplit)==4):                                  #Se valida que el asunto del correo tenga dos palabras
            if  (subjectSplit[0]=="Accept:" or  subjectSplit[0]=="Decline:"):
                SubjectsArray.append(subjectSplit[3])

    contlen=0
    ackArr=0
    contRef=0

    for index in range(len(SubjectsArray)):         
        for referencia in SubjectsArray:
            if contlen<len(SubjectsArray):
                if SubjectsArray[index]==' ':
                    pass
                elif SubjectsArray[index]==referencia:
                    contRef+=1
                    rep=referencia
                contlen+=1
        if contRef>1:
            repetidosArr.append(rep)
            numRepetidos.append(contRef) 
            for i in range(contRef):
                SubjectsArray.remove(rep)
                SubjectsArray.insert(0,' ')
        contRef=0
        contlen=0

    arrComplexData.append(repetidosArr)
    arrComplexData.append(numRepetidos)

    for message in messages:                                        #Hace un barrido por cada correo en bandeja
        
        subject_content = message.Subject
        subjectSplit=subject_content.split()                      #Separa el asunto del correo por espacios
        if (len(subjectSplit)==4):                                #Se valida que el asunto del correo tenga cuatro palabras  
            
            for r in range(len(arrComplexData[0])):
                if arrComplexData[0][r]==subjectSplit[3]:
                    repetido=arrComplexData[0][r]
                    repetidos=arrComplexData[1][r]
                    arrComplexData[0].pop(r)
                    arrComplexData[1].pop(r)
                    ackArr=1

            if ackArr==0:
                if (subjectSplit[0]=="Accept:"):                        #Condicion que determina si la tarifa fue aceptada   
                    respuestaPor=str(message.Sender)        
                    respSplit=respuestaPor.split()
                    respStr=''
                    print('row_count',row_count)
                    for i in range (0,3):
                        respStr+=respSplit[i]+' '
                    for i in range(row_count-1):                        #Ciclo que barre la base de datos csv en busca de la tarifa encontrada en el correo
                        if (baseDatos['Referencia'][i]==int(subjectSplit[3])):           #Se busca el match de la referencia del correo con alguno de la base de datos
                            if(baseDatos['Respuesta'][i]=='Aceptado' or baseDatos['Respuesta'][i]=='Rechazado'):
                                pass
                            else:
                                print(baseDatos['Referencia'][i])
                                sheetName.cell(i+2,11).value=datetime.datetime.now()              #Se escri+1be la hora de respuesta
                                sheetName.cell(i+2,10).value='Aceptado' 
                                sheetName.cell(i+2,16).value=respStr                    #Se cambia el valor de la columna 'Respuesta' de la referecia en cuestión a 'Aceptado'

                if (subjectSplit[0]=="Decline:"):                       #Mismo proceso pero para rechazados
                    respuestaPor=str(message.Sender)
                    respSplit=respuestaPor.split()
                    respStr=''
                    for i in range (0,3):
                        respStr+=respSplit[i]+' '
                        for i in range(row_count-1):
                            if (baseDatos['Referencia'][i]==int(subjectSplit[3])):
                                if(baseDatos['Respuesta'][i]=='Rechazado' or baseDatos['Respuesta'][i]=='Aceptado'):
                                    pass
                                else:
                                    sheetName.cell(i+2,11).value=datetime.datetime.now()               #Se escribe la hora de respuesta
                                    sheetName.cell(i+2,10).value='Rechazado' 
                                    sheetName.cell(i+2,16).value=respStr                    #Se cambia el valor de la columna 'Respuesta' de la referecia en cuestión a 'Aceptado'
            elif ackArr==1:
                contador=0
                for mess in messages:
                    subject_contentR = mess.Subject
                    subjectSplitR=subject_contentR.split()
                    if (len(subjectSplitR)==4):  
                        if  (subjectSplitR[0]=="Accept:" or  subjectSplitR[0]=="Decline:"):
                            if subjectSplitR[3]==repetido:
                                print(subjectSplitR[3])
                                contador+=1
                                print (contador)
                                if contador!=repetidos:
                                    pass
                                elif contador==repetidos:
                                    if (subjectSplitR[0]=="Accept:"):                        #Condicion que determina si la tarifa fue aceptada   
                                        respuestaPor=str(mess.Sender)        
                                        respSplit=respuestaPor.split()
                                        respStr=''
                                        for i in range (0,3):
                                            respStr+=respSplit[i]+' '
                                        for i in range(row_count-1):                        #Ciclo que barre la base de datos csv en busca de la tarifa encontrada en el correo
                                            if (baseDatos['Referencia'][i]==int(subjectSplitR[3])):           #Se busca el match de la referencia del correo con alguno de la base de datos
                                                if(baseDatos['Respuesta'][i]=='Aceptado' or baseDatos['Respuesta'][i]=='Rechazado'):
                                                    pass
                                                else:
                                                    sheetName.cell(i+2,11).value=datetime.datetime.now()              #Se escri+1be la hora de respuesta
                                                    sheetName.cell(i+2,10).value='Aceptado' 
                                                    sheetName.cell(i+2,16).value=respStr                    #Se cambia el valor de la columna 'Respuesta' de la referecia en cuestión a 'Aceptado'

                                    if (subjectSplitR[0]=="Decline:"):                       #Mismo proceso pero para rechazados
                                        respuestaPor=str(mess.Sender)
                                        respSplit=respuestaPor.split()
                                        respStr=''
                                        for i in range (0,3):
                                            respStr+=respSplit[i]+' '
                                            for i in range(row_count-1):
                                                if (baseDatos['Referencia'][i]==int(subjectSplitR[3])):
                                                    if(baseDatos['Respuesta'][i]=='Rechazado' or baseDatos['Respuesta'][i]=='Aceptado'):
                                                        pass
                                                    else:
                                                        sheetName.cell(i+2,11).value=datetime.datetime.now()               #Se escribe la hora de respuesta
                                                        sheetName.cell(i+2,10).value='Rechazado' 
                                                        sheetName.cell(i+2,16).value=respStr                    #Se cambia el valor de la columna 'Respuesta' de la referecia en cuestión a 'Aceptado'
                ackArr=0

    try:
        xfile.save(fname)     
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fname)
        ws = wb.Worksheets("Sheet1")
        ws.Columns.AutoFit()
        wb.Save()
        excel.Application.Quit() 
    except (PermissionError, TypeError) as w:
        print(w)
        print('El archivo se encuentra abierto. Espere.')
        time.sleep(1)
        while True:
            try:
                xfile.save(fname)
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(fname)
                ws = wb.Worksheets("Sheet1")
                ws.Columns.AutoFit()
                wb.Save()
                excel.Application.Quit()
                break
            except (PermissionError, TypeError) as w:
                print(w)
                print('El archivo sigue abierto. Espere.')  
                time.sleep(1) 

def correosAceptadosPorTiempo():           #Funcion que acepta correos por tiempo excedido
    if re.findall('~',arrFiles[-1]):
        fname=arrFiles[-2]
    else:
        fname=arrFiles[-1]
    workbook = xlrd.open_workbook(fname, on_demand=True)        #Determina el numero de filas
    sheet=workbook.sheet_by_name('Sheet1')
    row_count=sheet.nrows                   #Determina el numero de filas
    workbook.release_resources()
    baseDatos = pd.read_excel(fname)
    xfile = load_workbook(fname)
    sheetName = xfile.get_sheet_by_name('Sheet1') 
    for i in range(row_count-1):               #Barrido de datos en la base de datos en la columna 'Hora de envio'
        if (baseDatos['Hora de envio'][i])=='-':
            pass
        elif(baseDatos['Respuesta'][i])!='-':
            pass
        elif (((pd.to_datetime(baseDatos['Hora de envio'][i])+datetime.timedelta(minutes=20))<datetime.datetime.now())):      #Si la hora de envio excedió 30 mins,
            sheetName.cell(i+2,11).value=datetime.datetime.now()                                                 
            sheetName.cell(i+2,10).value='Aceptado por tiempo'                                                               #Se modifica la columna 'Respuesta' a 'Aceptado por tiempo'
    try:
        xfile.save(fname)  
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fname)
        ws = wb.Worksheets("Sheet1")
        ws.Columns.AutoFit()
        wb.Save()
        excel.Application.Quit()
    except (PermissionError, TypeError) as w:
        print(w)
        print('El archivo se encuentra abierto. Espere')
        time.sleep(1)
        while True:
            try:
                xfile.save(fname)  
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(fname)
                ws = wb.Worksheets("Sheet1")
                ws.Columns.AutoFit()
                wb.Save()
                excel.Application.Quit()
                break
            except (PermissionError, TypeError) as w:
                print(w)
                print('El archivo sigue abierto. Espere')
                time.sleep(1)

def proceso():
    succesfullRetrieval() #Se leen los correos en bandeja
    correosAceptadosPorTiempo()  #Se modifican los que exceden el tiempo de respuesta
    

schedule.every(25).minutes.do(proceso)

while True:
    schedule.run_pending()
    time.sleep(1)
