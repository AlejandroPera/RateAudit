import win32com.client as win32
import datetime
import pandas as pd
import easygui
import xlrd
from glob import glob
from openpyxl import load_workbook
import os.path
import xlsxwriter
import time
import re

path='C:\\Users\\aperalda\\Test\\'            #direccion donde se guarda el csv creado
arrFiles=glob(path+"*.xlsx")
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts
arrDest=[]
arrLin=[]
masArr=[]



def tipodeUnidad():
    msg ="Selecciona el tipo de unidad"
    title = "Tipo de unidad"
    choicesT = ["FULL",'THNC','53PSO','THNM','53T','THN','48TC','THNB','48TP','TN6C','48T','48TB','THNP','48M','TN6','48PSO','MOTO','MDZ','15TR','TN6P','TN6M','RBN','RBNB','RBNC','RBNM','RBNP','10T','12T','15C','THNR','48TR','RBNR','35TP','35TR','35T','TEST','30T','15TB','15T','35TC','35TB']
    choiceT = easygui.choicebox(msg, title, choicesT)
    if choiceT==None:     # show a Continue/Cancel dialog
            MensajeBienv()
    else:  # user chose Cancel
        if easygui.ccbox("Escogiste: {}".format(choiceT)):
            return choiceT
        else:
            tipodeUnidad()
            return choiceT

def cliente():
    msg ="Selecciona el cliente"
    title = "Cliente"
    choicesC = ['MXPLAN/MXAAG','MXPLAN/MXATT','MXPLAN/MXABT','MXPLAN/MXACC','MXPLAN/MXACH','MXPLAN/MXALA', 'MXPLAN/MXAME','MXPLAN/MXAPP','MXPLAN/MXASE','MXPLAN/MXATT','MXPLAN/MXBAR','MXPLAN/MXBEA','MXPLAN/MXBIO','MXPLAN/MXCAD','MXPLAN/MXCAL','MXPLAN/MXCAM','MXPLAN/MXCIB','MXPLAN/MXCOK','MXPLAN/MXCON','MXPLAN/MXCOP','MXPLAN/MXCOR','MXPLAN/MXCOT','MXPLAN/MXCUE','MXPLAN/MXDAR','MXPLAN/MXDEV','MXPLAN/MXEFC','MXPLAN/MXEFL','MXPLAN/MXEFM','MXPLAN/MXEFZ','MXPLAN/MXENG','MXPLAN/MXESA','MXPLAN/MXEVE','MXPLAN/MXFOD','MXPLAN/MXGAT','MXPLAN/MXGLX','MXPLAN/MXGMI','MXPLAN/MXHAM','MXPLAN/MXHBS','MXPLAN/MXHER','MXPLAN/MXHPC','MXPLAN/MXHPE','MXPLAN/MXIDV','MXPLAN/MXIMD','MXPLAN/MXIMP','MXPLAN/MXIMX','MXPLAN/MXIMX','MXPLAN/MXISS','MXPLAN/MXJAN','MXPLAN/MXJCT','MXPLAN/MXJNJ','MXPLAN/MXKAZ','MXPLAN/MXKEL','MXPLAN/MXKFT','MXPLAN/MXKIT','MXPLAN/MXLEG','MXPLAN/MXLEN','MXPLAN/MXLGE','MXPLAN/MXMAP','MXPLAN/MXMAT','MXPLAN/MXMAV','MXPLAN/MXMER','MXPLAN/MXMOK','MXPLAN/MXNOV','MXPLAN/MXNXT','MXPLAN/MXOGL','MXPLAN/MXPIL','MXPLAN/MXPIL','MXPLAN/MXPLAN','MXPLAN/MXPNG','MXPLAN/MXPSO','MXPLAN/MXRYC','MXPLAN/MXSBS','MXPLAN/MXSDK','MXPLAN/MXSHW','MXPLAN/MXSOL','MXPLAN/MXSON','MXPLAN/MXSUN','MXPLAN/MXSWM','MXPLAN/MXTAJ','MXPLAN/MXTCT','MXPLAN/MXTCY','MXPLAN/MXTEV','MXPLAN/MXTRN','MXPLAN/MXTYC','MXPLAN/MXUHU','MXPLAN/MXUNI','MXPLAN/MXWKT','MXPLAN/MXZUU']
    choiceC = easygui.choicebox(msg, title, choicesC)
    if choiceC==None:     # show a Continue/Cancel dialog
            tipodeUnidad()
    else:  # user chose Cancel
        conf=easygui.ccbox("Escogiste: {}".format(choiceC))
        if conf:
            return choiceC
        else:
            cliente()
            return choiceC

def origen():
    msg ="Selecciona el origen"
    title = "Origen"
    choicesO = ['37D002','37D004','37D005','37D009','37D014','37D015','37D016','37D019','37D024','37D027','37D031','37D035','37D037','37D050','37D051','37D065','37D079','37D082','37D090','37D091','37D092','37D100','37D110','37D112','37D115','37D116','37D117','37D118','37D120','37D121','37D122','37D123','37D124','37D128','37D129','37D130','37D131','37D132','37D133','37D138','37D139','37D140','37D142','37D143','37D145','37D146','37D147','37D151']
    choiceO = easygui.choicebox(msg, title, choicesO)
    if choiceO==None:     # show a Continue/Cancel dialog
            cliente()
    else:  # user chose Cancel
        conf=easygui.ccbox("Escogiste: {}".format(choiceO))
        if conf:
            return choiceO
        else:
            origen()
            return choiceO

def bucleConfirmacion(fValues, valoresPick):
    msg = "Ingresa los datos solicitados"
    title = "Control de correos"
    fieldNames = ["CV","Destino","Tarifa",'Linea','Numero de referencia','Correo','CC (no es obligatorio)','CC2 (no es obligatorio)','CC3 (no es obligatorio)']
    fieldValues = easygui.multenterbox(msg, title, fieldNames,fValues)
    cont=0
    ack=0
    errmsgF=''
    print(fieldValues)
    while 1:
        while cont<=5:
            if fieldValues[cont].strip() == "":
                errmsgF += "{} is a required field.\n\n".format(fieldNamees[cont])
                ack=1
            cont+=1
        if ack==1:
            while 1:
                if errmsgF=='':
                    break
                fieldValues = easygui.multenterbox(errmsgF, title, fieldNames, fieldValues)
                errmsgF=''
                for b in range(0,6):
                    print(fieldValues[b])
                    print(fieldNamees[b])
                    if fieldValues[b].strip() == "":
                        print(fieldValues[b])
                        errmsgF+="{} is a required field.\n\n".format(fieldNamees[b])
                        print(errmsgF)
        print(fieldValues)
        parseD=fieldValues[1]
        parseL=fieldValues[3]
        for i in parseD:
            arrDest.append(i)
        for i in parseL:
            arrLin.append(i)
        while (arrDest[0].upper()!='M'or arrDest.upper()[1]!='X'):
            for i in range(len(arrDest)):
                arrDest.pop(0)  
            errmsgD= "{}: Usa formato MX/#.\n\n".format(fieldNames[1])
            fieldValues = easygui.multenterbox(errmsgD, title, fieldNames, fieldValues)
            parseDNew=fieldValues[1]
            for i in parseDNew:
                arrDest.append(i)
        while (arrLin[0].upper()!='M'or arrLin[1].upper()!='X'):
            for i in range(len(arrLin)):
                arrLin.pop(0)
            errmsgL= "{}: Usa formato MX/#.\n\n".format(fieldNames[3])
            fieldValues = easygui.multenterbox(errmsgL, title, fieldNames, fieldValues)
            parseLNew=fieldValues[3]
            for i in parseLNew:
                arrLin.append(i)
        break
    print("Reply was:{}".format(fieldValues))
    confirm=confirmacion(fieldValues, valoresPick)
    print(confirm,'Bucle')
    return confirm

def confirmacion(fieldValues, valoresPick):
    msg = 'CV: '+fieldValues[0]+'\n'+'Cliente: '+valoresPick[1]+'\n'+'Destino: '+fieldValues[1]+'\n'+'Origen: '+valoresPick[2]+'\n'+'Tipo de Unidad: '+valoresPick[0]+'\n'+'Tarifa: '+fieldValues[2]+'\n'+'Linea: '+fieldValues[3]+'\n'+'Numero de referencia: '+fieldValues[4]+'\n'+'Correo: ' +fieldValues[5]
    title = "Prompt de confrmación"
    if easygui.ccbox(msg, title):     # show a Continue/Cancel dialog
        pass  # user chose Continue
        return fieldValues
    else:  # user chose Cancel
        bucleData=bucleConfirmacion(fieldValues,valoresPick)
        print(bucleData,'bucleData')
        return bucleData
        
def start():
    tipo=tipodeUnidad()
    cli=cliente()
    ori=origen()
    #'cliente','origen','tipo de unidad'
    msg = "Ingresa los datos solicitados"
    title = "Control de correos"
    fieldNames = ["CV","Destino (MX/#)","Tarifa",'Linea (MX/#)','Numero de referencia','Correo','CC (no es obligatorio)','CC2 (no es obligatorio)','CC3(no es obligatorio)']
    fieldNamees = ["CV","Destino (MX/#)","Tarifa",'Linea (MX/#)','Numero de referencia','Correo']
    fieldValues = easygui.multenterbox(msg, title, fieldNames)
    while fieldValues==None:
        fieldValues = easygui.multenterbox(msg, title, fieldNames)
    cont=0
    ack=0
    errmsgF=''
    print(fieldValues)
    while 1:
        while cont<=5:
            if fieldValues[cont].strip() == "":
                errmsgF += "{} is a required field.\n\n".format(fieldNamees[cont])
                ack=1
            cont+=1
        if ack==1:
            while 1:
                if errmsgF=='':
                    break
                fieldValues = easygui.multenterbox(errmsgF, title, fieldNames, fieldValues)
                errmsgF=''
                for b in range(0,6):
                    print(fieldValues[b])
                    print(fieldNamees[b])
                    if fieldValues[b].strip() == "":
                        print(fieldValues[b])
                        errmsgF+="{} is a required field.\n\n".format(fieldNamees[b])
                        print(errmsgF)
        print(fieldValues)
        parseD=fieldValues[1]
        parseL=fieldValues[3]
        for i in parseD:
            arrDest.append(i)
        for i in parseL:
            arrLin.append(i)
        while (arrDest[0].upper()!='M'or arrDest[1].upper()!='X'):
            for i in range(len(arrDest)):
                arrDest.pop(0)  
            errmsgD= "{}: Usa formato MX/#.\n\n".format(fieldNames[1])
            fieldValues = easygui.multenterbox(errmsgD, title, fieldNames, fieldValues)
            parseDNew=fieldValues[1]
            for i in parseDNew:
                arrDest.append(i)
        while (arrLin[0].upper()!='M'or arrLin[1].upper()!='X'):
            for i in range(len(arrLin)):
                arrLin.pop(0)
            errmsgL= "{}: Usa formato MX/#.\n\n".format(fieldNames[3])
            fieldValues = easygui.multenterbox(errmsgL, title, fieldNames, fieldValues)
            parseLNew=fieldValues[3]
            for i in parseLNew:
                arrLin.append(i)
        break
    valoresPick=[tipo,cli,ori]
    datosConf=confirmacion(fieldValues, valoresPick)
    datosConf.insert(1,cli)
    datosConf.insert(3,ori)
    datosConf.insert(4,tipo)
    print('datos af', datosConf)
    return datosConf                                                 #Return del arreglo


def succesfullSend(data):                      #Funcion que manda el mail. Recibe la dirección de correo, datos necesarios p/registro, num de referencia y el nombre del archivo a crear (csv)
    outlook = win32.Dispatch('outlook.application')
    xfile = load_workbook(fname)
    sheetName = xfile.get_sheet_by_name('Sheet1')
    workbook = xlrd.open_workbook(fname)        #Determina el numero de filas
    sheet=workbook.sheet_by_name('Sheet1')
    row_count=int(sheet.nrows) 
    print(range(len((data))))
    for i in range(len((data))):
        cv=data[i][1]
        cliente=data[i][2]
        Destino=data[i][3]
        Origen=data[i][4]
        TipodeU=data[i][5]
        tarifa=data[i][6]
        linea=data[i][7]
        email_data = [cv,cliente,Destino,Origen,TipodeU,tarifa,linea]
        mail = outlook.CreateItem(0) 
        cadena_destinatarios = data[i][11]
        if data[i][12]!='':
            cadena_destinatarios += ';' + data[i][12]
        if data[i][13]!='':
            cadena_destinatarios += ';' + data[i][13]
        if data[i][14]!='':
            cadena_destinatarios += ';' + data[i][14]

        mail.To = cadena_destinatarios
        mail.VotingOptions = "Accept;Decline"                       #Activate voting options
        mail.Subject = 'Autorización spot: '+str(data[i][0])                 #Num de referencia se agrega al asunto del correo
        body = txt_to_str("mail_format.html")
        for i in range(len(email_data)):
            body = body.replace("<!--SPLITCONTROL("+str(i)+")-->","<td>{0}</td>".format(str(email_data[i])))        
        mail.HTMLBody = body
        images_path = "C:\\Users\\aperalda\\Documents\\RAudit\\RateAudit\\img\\"
        mail.Attachments.Add(Source= images_path+"voting_buttons.png")
        mail.Attachments.Add(Source= images_path+"llpc.PNG")                                 #Los datos se agregan al cuerpo del correo   
        print(row_count)
        mail.Send()                                                 #Se manda
        #Agregar condición de verificar envío 
        print("Successful send")                                    #Wuuuuu
        for cont in range(len(data)):
            for c in range(2,row_count+1):
                print(sheetName.cell(c,1).value)
                print(data[cont][0])
                if sheetName.cell(c,1).value==data[cont][0]:
                    sheetName.cell(c,9).value=datetime.datetime.now()
        try:
            xfile.save(fname)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            ws = wb.Worksheets("Sheet1")
            ws.Columns.AutoFit()
            wb.Save()
            excel.Application.Quit()
        except (PermissionError, TypeError) as w:
            print('Espere un momento. El archivo está abierto')
            time.sleep(1)
            while True:
                try:
                    xfile.save(fname)
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    wb = excel.Workbooks.Open(fname)
                    ws = wb.Worksheets("Sheet1")
                    ws.Columns.AutoFit()
                    wb.Save()
                    excel.Application.Quit()
                    break
                except (PermissionError, TypeError) as w:
                    time.sleep(1)
                    print('Espere un momento. El archivo sigue abierto')
                    cont=0
        

def txt_to_str(route):
    f = open(route, mode="r", encoding="utf-8")
    content = f.read()
    f.close()
    return str(content)
                                         

def creaCSV(datos):                               #Funcion que crea la base de datos csv
    #if (datos[0]=='' or datos[1]=='' or datos[2]=='' or datos[3]=='' or datos[4]=='' or datos[5]=='' or datos[6]=='' or datos[7]=='' or datos[8]=='' or datos[9]==''):
    cv=datos[0]
    cliente=datos[1]
    destino=datos[2]
    origen=datos[3]
    tipodeu=datos[4]
    tarifa=datos[5]
    linea=datos[6]
    referencia=datos[7]
    correo=datos[8]
    print(datos)
    for i in range(len(datos)-1):
        if len(datos[i])<2:
            datos.pop(i)
    if len(datos[-1])<3:
        datos.pop(-1)
    print(datos)
    if len(datos)<10:
        correo1=""
        correo2=""
        correo3=""
        print('Len9')
    elif len(datos)<11:
        correo1=datos[9]
        correo2=""
        correo3=""
        print('Len10')
    elif len(datos)<12:


        correo1=datos[9]
        correo2=datos[10]
        correo3=""
        print('Len11')
    else:
        correo1=datos[9]
        correo2=datos[10]
        correo3=datos[11]
        print('Len12')

    addToInit=[referencia,cv, cliente,destino,origen,tipodeu,tarifa,linea,'-','-','-',correo,correo1,correo2,correo3,'-']   #Se agregan los valores obtenidos al csv con la columna de respuesta con valor '-'
    # Se abre el archivo en modo append
    workbook = xlrd.open_workbook(fname)        #Determina el numero de filas
    sheet=workbook.sheet_by_name('Sheet1')
    row=sheet.nrows
    col=1
    xfile = load_workbook(fname)
    sheetName = xfile.get_sheet_by_name('Sheet1')
    for i in addToInit:
        sheetName.cell(row+1,col).value=i
        col=col+1
    try:
        xfile.save(fname)
    except PermissionError:
        print('Espere un momento. Archivo abierto')
        time.sleep(1)
        while True:
            try:
                xfile.save(fname)
                break
            except PermissionError:
                time.sleep(1)
                print('El archivo sigue abierto. Espere.')
    return addToInit 
    
def nombreArchivo():                                            #Funcion que crea el nombre del archivo
    nowC=datetime.datetime.now()
    mes=nowC.strftime("%B")
    anho=nowC.strftime("%Y")
    fullName=mes+'-'+anho+'.xlsx'
    fullStr=path+fullName
    df = pd.DataFrame({'Referencia':[''], 'CV':[''],'Cliente':[''],'Destino':[''],'Origen':[''],'Tipo de Unidad':[''],'Tarifa':[''],'Linea':[''],'Hora de envio':[''],'Respuesta':[''],'Hora de respuesta':[''],'Correo':[''],'CC1':[''],'CC2':[''],'CC3':[''],'Persona que responde':['']})
    writer = pd.ExcelWriter(fullStr, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', startrow=1, header=False, index=False)
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']
    header_format = workbook.add_format({
    'bold': True,
    'align': 'center',
    'fg_color': '#FFFF00',
    'border_color': '#000000',
    'border':1
    })
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    writer.save()
    # Close the Pandas Excel writer and output the Excel file.
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fullStr)
    ws = wb.Worksheets("Sheet1")
    ws.Columns.AutoFit()
    wb.Save()
    excel.Application.Quit()
    return fullStr

def checkArchivo():
    nowC=datetime.datetime.now()
    mes=nowC.strftime("%B")
    anho=nowC.strftime("%Y")
    fullName=mes+'-'+anho+'.xlsx'
    fullname=path+fullName
    return fullname


def MensajeBienv():
    Bienvenido=easygui.ccbox(msg='Bienvenido al sistema DHL', title='DHL Supply Chain', choices=('Agregar', 'Enviar'), image='img\\dhl_logo.png',default_choice='Agregar', cancel_choice='Enviar')
    if  Bienvenido==True:
            datos=start()                                                   #Se piden los datos
            arrDatos=creaCSV(datos)
            masArr.append(arrDatos)
            print(masArr)
            MensajeBienv()
    elif Bienvenido==False:
        if len(masArr)==0:
            print('Primero aregega antes de enviar')
            MensajeBienv()
        succesfullSend(masArr)                                         #Cuando exista valor en la variable fileName, se toma el tiempo actual

        
print(datetime.date.today().month)
txt_to_str("mail_format.html")
#Corregir, hacer patch
if len(arrFiles)>0:
    print(arrFiles)
    if re.findall('~',arrFiles[-1]):
        archivo=arrFiles[-2]
    else:
        archivo=arrFiles[-1]

    archivoSplit=archivo.split('\\')
    nomenclaturaArchivo=archivoSplit[-1]
    pointMes=nomenclaturaArchivo.split('-')
    checkMes=pointMes[0]
    print(checkMes)
    if checkMes=='January':
        varMes=1
    elif checkMes=='February':
        varMes=2
    elif checkMes=='March':
        varMes=3
    elif checkMes=='April':
        varMes=4
    elif checkMes=='May':
        varMes=5
    elif checkMes=='June':
        varMes=6
    elif checkMes=='July':
        varMes=7
    elif checkMes=='August':
        varMes=8
    elif checkMes=='September':
        varMes=9
    elif checkMes=='October':
        varMes=10
    elif checkMes=='November':
        varMes=11
    elif checkMes=='December':
        varMes=12
    if varMes==datetime.date.today().month:
        fname=checkArchivo()
    else:
        fname=nombreArchivo()                                            #Se crea el nombre del archivo
else:
    fname=nombreArchivo()


MensajeBienv()   